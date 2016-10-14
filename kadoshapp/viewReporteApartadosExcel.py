#Vista genérica para mostrar resultados
from django.views.generic.base import TemplateView
#Workbook nos permite crear libros en excel
from openpyxl import Workbook
from openpyxl.styles import colors
from openpyxl.styles import Font, Color,Fill,PatternFill
from openpyxl.cell import Cell
from openpyxl.worksheet.properties import WorksheetProperties, PageSetupProperties
#Nos devuelve un objeto resultado, en este caso un archivo de excel
from django.http.response import HttpResponse
from .models import *
from django.db.models import Sum
from django.db.models import Q
#El siguiente método convierte el resultado de "values" en un diccionario
def ValuesQuerySetToDict(vqs):
    return [item for item in vqs]
#Nuestra clase hereda de la vista genérica TemplateView

class ReporteApartados(TemplateView):
    def get(self, request, *args, **kwargs):
        #recibo los datos
        reportedeudas=Venta.objects.filter(estado_venta=1,es_cotizacion=0,contado_venta=0).values('pk','cliente_idcliente__persona_idpersona__nombres_persona','cliente_idcliente__persona_idpersona__apellidos_persona','total_venta','cuentaporcobrar__pk','cuentaporcobrar__saldo_inicial_cuentaporcobrar','cuentaporcobrar__saldo_actual_cuentaporcobrar')
        reporteApartado=ValuesQuerySetToDict(reportedeudas)
        #Creamos el libro de trabajo
        wb = Workbook()
        #Definimos como nuestra hoja de trabajo, la hoja activa, por defecto la primera del libro
        ws = wb.active
        #En la celda B1 ponemos el texto 'REPORTE DE PERSONAS'
        ws['B1'] = 'Kadosh'
        ws['B2'] = 'REPORTE DE APARTADOS'
        import datetime
        ws['B3'] = datetime.datetime.now()

        #Juntamos las celdas desde la B1 hasta la E1, formando una sola celda
        ws.merge_cells('B1:E1')
        ws.merge_cells('B2:E2')
        ws.merge_cells('B3:E3')
        #Colores de fuente de celda y tama;os de letras
        ft = Font(color=colors.RED)
        b1 = ws['B1']
        b1.font = ft
        ft2 = Font(color=colors.BLUE)
        b2 = ws['B2']
        b2.font = ft2

        #Creamos los encabezados desde la celda B3 hasta la E3
        ws['A5'] = 'Cod Venta'
        ws['B5'] = 'Nombres Cliente'
        ws['C5'] = 'Apellidos Cliente'
        ws['D5'] = 'Total Venta'
        ws['E5'] = 'Cod Cuenta'
        ws['F5'] = 'Saldo Inicial'
        ws['G5'] = 'Saldo Actual'

        cont=6
        #Recorremos el conjunto de personas y vamos escribiendo cada uno de los datos en las celdas
        for producto in reporteApartado:
            ws.cell(row=cont,column=1).value = producto['pk']
            ws.cell(row=cont,column=2).value = producto['cliente_idcliente__persona_idpersona__nombres_persona']
            ws.cell(row=cont,column=3).value = producto['cliente_idcliente__persona_idpersona__apellidos_persona']
            ws.cell(row=cont,column=4).value = producto['total_venta']
            ws.cell(row=cont,column=5).value = producto['cuentaporcobrar__pk']
            ws.cell(row=cont,column=6).value = producto['cuentaporcobrar__saldo_inicial_cuentaporcobrar']
            ws.cell(row=cont,column=7).value = producto['cuentaporcobrar__saldo_actual_cuentaporcobrar']
            cont = cont + 1
        #Establecemos el nombre del archivo
        nombre_archivo ="ReporteApartados.xlsx"
        #Definimos que el tipo de respuesta a devolver es un archivo de microsoft excel
        response = HttpResponse(content_type="application/ms-excel")
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response
