#Vista genérica para mostrar resultados
from django.views.generic.base import TemplateView
#Workbook nos permite crear libros en excel
from openpyxl import Workbook
from openpyxl.styles import colors
from openpyxl.styles import Font, Color,Fill,PatternFill
from openpyxl.cell import Cell
from openpyxl.worksheet.properties import WorksheetProperties, PageSetupProperties
#Nos devuelve un objeto resultado, en este caso un archivo de excel

from django.http.response import HttpResponse
from .models import *
from django.db.models import Sum, Count
from django.utils import timezone
import datetime
from django.db import connection
import pytz #para poder hacer la suma de los campos
from django.db.models import Q #para poder usar el operador | que funciona como OR
from django.db.models import F
from datetime import timedelta

def ValuesQuerySetToDict(vqs):
    return [item for item in vqs]

def to_dict(instance):
    opts = instance._meta
    data = {}
    for f in opts.concrete_fields + opts.many_to_many:
        if isinstance(f, ManyToManyField):
            if instance.pk is None:
                data[f.name] = []
            else:
                data[f.name] = list(f.value_from_object(instance).values_list('pk', flat=True))
        else:
            data[f.name] = f.value_from_object(instance)
    return data

#Nuestra clase hereda de la vista genérica TemplateView
class ReporteVentas(TemplateView):

    def get(self, request, *args, **kwargs):
        #recibo los datos
        fechaini = self.request.GET.get('fechainicial_precio')
        fechafini = self.request.GET.get('fechafinal_precio')
        empleado = self.request.GET.get('empleado_idempleado')
        checo=str(self.request.GET.get('checkbo'))

        if not empleado:
            empleado=0

        truncate_date = connection.ops.date_trunc_sql('month', 'fecha_venta') #se obtiene solo el mes de la venta
        qs = Venta.objects.extra({'month':truncate_date}) # se añade un nuevo campo/columna que tendrá de nombre "month" y tendrá como datos "truncate_date" que en este caso es el mes
        fecha1=str(fechaini) #debe ser la fecha más pequeña
        fecha2=str(fechafini) #debe ser la fecha más grande
        fecha1_split=fecha1.split('/')
        fecha2_split=fecha2.split('/')

        ventas_vendedor=qs.values('month','vendedor_venta__persona_idpersona__nombres_persona','vendedor_venta__persona_idpersona__apellidos_persona').annotate(total_ventas=Sum('total_venta')).order_by('month','total_ventas')
        pormes=False
        if len(fecha1_split)>1 and len(fecha2_split)>1:
            fechainicial_real=datetime.datetime(int(fecha1_split[2]),int(fecha1_split[1]), int(fecha1_split[0]),0,0,0,tzinfo=pytz.UTC)
            fechafinal_real=datetime.datetime(int(fecha2_split[2]),int(fecha2_split[1]), int(fecha2_split[0]),23,59,59,tzinfo=pytz.UTC)
            fechainicial_real=fechainicial_real+datetime.timedelta(hours=6)
            fechafinal_real=fechafinal_real+datetime.timedelta(hours=6)
            if checo=="true":
                pormes=True
                ventas_vendedor = qs.filter(entregada_venta=1,es_cotizacion=0,estado_venta=1,vendedor_venta=int(empleado),fecha_venta__range=(fechainicial_real, fechafinal_real)).values('month','vendedor_venta__persona_idpersona__nombres_persona','vendedor_venta__persona_idpersona__apellidos_persona').annotate(total_ventas=Sum('total_venta')).order_by('month') #este y funciona con las horas
                if not ventas_vendedor:
                    ventas_vendedor = qs.filter(entregada_venta=1,es_cotizacion=0,estado_venta=1,fecha_venta__range=(fechainicial_real,fechafinal_real)).values('month','vendedor_venta__persona_idpersona__nombres_persona','vendedor_venta__persona_idpersona__apellidos_persona').annotate(total_ventas=Sum('total_venta')).order_by('month') #este y funciona con las horas
            else:
                ventas_vendedor = Venta.objects.filter(entregada_venta=1,es_cotizacion=0,estado_venta=1,vendedor_venta=int(empleado),fecha_venta__range=(fechainicial_real, fechafinal_real)).values('vendedor_venta__persona_idpersona__nombres_persona','vendedor_venta__persona_idpersona__apellidos_persona').annotate(total_ventas=Sum('total_venta')).order_by('total_ventas')
                if not ventas_vendedor:
                    ventas_vendedor = Venta.objects.filter(entregada_venta=1,es_cotizacion=0,estado_venta=1,fecha_venta__range=(fechainicial_real, fechafinal_real)).values('vendedor_venta__persona_idpersona__nombres_persona','vendedor_venta__persona_idpersona__apellidos_persona').annotate(total_ventas=Sum('total_venta')).order_by('total_ventas')

        if not ventas_vendedor:
            ventas_vendedor=qs.values('month','vendedor_venta__persona_idpersona__nombres_persona','vendedor_venta__persona_idpersona__apellidos_persona').annotate(total_ventas=Sum('total_venta')).order_by('month','total_ventas')

        repo_ventas=ValuesQuerySetToDict(ventas_vendedor)

        #Creamos el libro de trabajo
        wb = Workbook()
        #Definimos como nuestra hoja de trabajo, la hoja activa, por defecto la primera del libro
        ws = wb.active
        #En la celda B1 ponemos el texto 'REPORTE DE PERSONAS'
        ws['B1'] = 'Kadosh'
        ws['B2'] = 'REPORTE DE VENTAS POR VENDEDOR'
        ws['B3'] = datetime.datetime.now()

        #Juntamos las celdas desde la B1 hasta la E1, formando una sola celda
        ws.merge_cells('B1:E1')
        ws.merge_cells('B2:E2')
        ws.merge_cells('B3:E3')
        #Colores de fuente de celda y tama;os de letras
        ft = Font(color=colors.RED)
        b1 = ws['B1']
        b1.font = ft
        ft2 = Font(color=colors.BLUE)
        b2 = ws['B2']
        b2.font = ft2

        #Creamos los encabezados desde la celda B3 hasta la E3
        ws['B5'] = 'Nombres'
        ws['C5'] = 'Apellidos'
        ws['D5'] = 'Mes'
        ws['E5'] = 'Venta'


        cont=6
        #Recorremos el conjunto de personas y vamos escribiendo cada uno de los datos en las celdas
        for vent in repo_ventas:
            if pormes:
                fecha=vent['month'].month #strftime('%B')
                ws.cell(row=cont,column=4).value = fecha
            ws.cell(row=cont,column=2).value = vent['vendedor_venta__persona_idpersona__nombres_persona']
            ws.cell(row=cont,column=3).value = vent['vendedor_venta__persona_idpersona__apellidos_persona']
            ws.cell(row=cont,column=5).value = vent['total_ventas']

            cont = cont + 1

        #Establecemos el nombre del archivo
        nombre_archivo ="ReporteVentasVendedor.xlsx"
        #Definimos que el tipo de respuesta a devolver es un archivo de microsoft excel
        response = HttpResponse(content_type="application/ms-excel")
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response
