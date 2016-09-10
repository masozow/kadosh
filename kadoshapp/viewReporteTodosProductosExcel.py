#Vista genérica para mostrar resultados
from django.views.generic.base import TemplateView
#Workbook nos permite crear libros en excel
from openpyxl import Workbook
from openpyxl.styles import colors
from openpyxl.styles import Font, Color,Fill,PatternFill
from openpyxl.cell import Cell
from openpyxl.worksheet.properties import WorksheetProperties, PageSetupProperties
#Nos devuelve un objeto resultado, en este caso un archivo de excel
from django.http.response import HttpResponse
from .models import *
from django.db.models import Sum
from django.db.models import Q
#El siguiente método convierte el resultado de "values" en un diccionario
def ValuesQuerySetToDict(vqs):
    return [item for item in vqs]
#Nuestra clase hereda de la vista genérica TemplateView

class ReporteProductos(TemplateView):
    def get(self, request, *args, **kwargs):
        #recibo los datos
        resp_producto=Producto.objects.filter(estado_producto=1).values('pk','nombre_producto','codigobarras_producto','codigoestilo_producto','marca_id_marca__nombre_marca','tipo_producto_idtipo_producto__nombre_tipoproducto','talla_idtalla__nombre_talla','color_idcolor__nombre_color','genero_idgener__nombre_genero','estilo_idestilo__nombre_estilo','inventarioproducto__bodega_idbodega__nombre_bodega','inventarioproducto__existencia_actual')
        txt_codigo_barras = request.GET.get('codigobarras_producto')
        if txt_codigo_barras:
            resp_producto=resp_producto.filter(codigobarras_producto=txt_codigo_barras)
        else:
            txt_codigo_producto = request.GET.get('codigoestilo_producto')
            if txt_codigo_producto:
                resp_producto=resp_producto.filter(codigoestilo_producto=txt_codigo_producto)
            
            id_marca_producto = request.GET.get('marca_id_marca')
            if id_marca_producto:
                resp_producto=resp_producto.filter(marca_id_marca=id_marca_producto)

            id_estilo_producto = request.GET.get('estilo_idestilo')
            if id_estilo_producto:
                resp_producto=resp_producto.filter(estilo_idestilo=id_estilo_producto)

            id_tipo_producto = request.GET.get('tipo_producto_idtipo_producto')
            if id_tipo_producto:
                resp_producto=resp_producto.filter(tipo_producto_idtipo_producto=id_tipo_producto)

            id_talla_producto = request.GET.get('talla_idtalla')
            if id_talla_producto:
                resp_producto=resp_producto.filter(talla_idtalla=id_talla_producto)

            id_color_producto = request.GET.get('color_idcolor')
            if id_color_producto:
                resp_producto=resp_producto.filter(color_idcolor=id_color_producto)

            id_genero_producto = request.GET.get('genero_idgener')
            if id_genero_producto:
                resp_producto=resp_producto.filter(genero_idgener=id_genero_producto)


        #productos=Producto.objects.filter(Q(codigobarras_producto=codigobarras)|Q(codigoestilo_producto=codigoEstilo) | Q(marca_id_marca=marca_int) | Q(estilo_idestilo=estilo_int )| Q(tipo_producto_idtipo_producto=tipo_int) | Q(talla_idtalla=talla_int) | Q(color_idcolor=color_int) | Q(genero_idgener=genero_int),estado_producto=1).values('pk','nombre_producto','codigobarras_producto','codigoestilo_producto','marca_id_marca__nombre_marca','tipo_producto_idtipo_producto__nombre_tipoproducto','talla_idtalla__nombre_talla','color_idcolor__nombre_color','genero_idgener__nombre_genero','estilo_idestilo__nombre_estilo','inventarioproducto__bodega_idbodega__nombre_bodega','inventarioproducto__existencia_actual')
        #if not productos:
        #    productos=Producto.objects.filter(estado_producto=1).values('pk','nombre_producto','codigobarras_producto','codigoestilo_producto','marca_id_marca__nombre_marca','tipo_producto_idtipo_producto__nombre_tipoproducto','talla_idtalla__nombre_talla','color_idcolor__nombre_color','genero_idgener__nombre_genero','estilo_idestilo__nombre_estilo','inventarioproducto__bodega_idbodega__nombre_bodega','inventarioproducto__existencia_actual')

        nuevos_productos=ValuesQuerySetToDict(resp_producto)
        #Creamos el libro de trabajo
        wb = Workbook()
        #Definimos como nuestra hoja de trabajo, la hoja activa, por defecto la primera del libro
        ws = wb.active
        #En la celda B1 ponemos el texto 'REPORTE DE PERSONAS'
        ws['B1'] = 'Kadosh'
        ws['B2'] = 'REPORTE EXISTENCIAS DE PRODUCTOS'
        import datetime
        ws['B3'] = datetime.datetime.now()

        #Juntamos las celdas desde la B1 hasta la E1, formando una sola celda
        ws.merge_cells('B1:E1')
        ws.merge_cells('B2:E2')
        ws.merge_cells('B3:E3')
        #Colores de fuente de celda y tama;os de letras
        ft = Font(color=colors.RED)
        b1 = ws['B1']
        b1.font = ft
        ft2 = Font(color=colors.BLUE)
        b2 = ws['B2']
        b2.font = ft2

        #Creamos los encabezados desde la celda B3 hasta la E3
        ws['A5'] = 'Cod'
        ws['B5'] = 'Producto'
        ws['C5'] = 'CodBarras'
        ws['D5'] = 'CodEstilo'
        #ws['F5'] = 'Precio'
        ws['E5'] = 'Marca'
        ws['F5'] = 'Tipo'
        ws['G5'] = 'Talla'
        ws['H5'] = 'Color'
        ws['I5'] = 'Genero'
        ws['J5'] = 'Estilo'
        ws['K5'] = 'Bodega'
        ws['L5'] = 'Existencias'

        cont=6
        #Recorremos el conjunto de personas y vamos escribiendo cada uno de los datos en las celdas
        for producto in nuevos_productos:
            ws.cell(row=cont,column=1).value = producto['pk']
            ws.cell(row=cont,column=2).value = producto['nombre_producto']
            ws.cell(row=cont,column=3).value = producto['codigobarras_producto']
            ws.cell(row=cont,column=4).value = producto['codigoestilo_producto']
            #ws.cell(row=cont,column=6).value = producto['precio__valor_precio']
            ws.cell(row=cont,column=5).value = producto['marca_id_marca__nombre_marca']
            ws.cell(row=cont,column=6).value = producto['tipo_producto_idtipo_producto__nombre_tipoproducto']
            ws.cell(row=cont,column=7).value = producto['talla_idtalla__nombre_talla']
            ws.cell(row=cont,column=8).value = producto['color_idcolor__nombre_color']
            ws.cell(row=cont,column=9).value = producto['genero_idgener__nombre_genero']
            ws.cell(row=cont,column=10).value = producto['estilo_idestilo__nombre_estilo']
            ws.cell(row=cont,column=11).value = producto['inventarioproducto__bodega_idbodega__nombre_bodega']
            ws.cell(row=cont,column=12).value = producto['inventarioproducto__existencia_actual']
            cont = cont + 1
        #Establecemos el nombre del archivo
        nombre_archivo ="ReporteExistenciasProductos.xlsx"
        #Definimos que el tipo de respuesta a devolver es un archivo de microsoft excel
        response = HttpResponse(content_type="application/ms-excel")
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response
