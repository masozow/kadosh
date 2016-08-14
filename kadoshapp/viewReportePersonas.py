#Vista genérica para mostrar resultados
from django.views.generic.base import TemplateView
#Workbook nos permite crear libros en excel
from openpyxl import Workbook
from openpyxl.styles import colors
from openpyxl.styles import Font, Color,Fill
from openpyxl.worksheet.properties import WorksheetProperties, PageSetupProperties
#Nos devuelve un objeto resultado, en este caso un archivo de excel
from django.http.response import HttpResponse
from django.views.generic.list import ListView
from .models import *
class Personas(ListView):
    model = Persona
    template_name = 'personas.html'
    context_object_name = 'personas'

#Nuestra clase hereda de la vista genérica TemplateView
class ReportePersonas(TemplateView):

    #Usamos el método get para generar el archivo excel
    def get(self, request, *args, **kwargs):
        #Obtenemos todas las personas de nuestra base de datos
        personas = Persona.objects.all()
        #Creamos el libro de trabajo
        wb = Workbook()
        #Definimos como nuestra hoja de trabajo, la hoja activa, por defecto la primera del libro
        ws = wb.active

        #En la celda B1 ponemos el texto 'REPORTE DE PERSONAS'
        ws['B1'] = 'Kadosh'
        ws['B2'] = 'REPORTE DE PERSONAS'
        ws['F1'] = 'Fecha'
        #Juntamos las celdas desde la B1 hasta la E1, formando una sola celda
        ws.merge_cells('B1:E1')
        ws.merge_cells('B2:E2')

        #Colores de fuente de celda
        ft = Font(color=colors.RED)
        b1 = ws['B1']
        b1.font = ft


        #Creamos los encabezados desde la celda B3 hasta la E3
        ws['B4'] = 'ID'
        ws['C4'] = 'NOMBRES'
        ws['D4'] = 'APELLIDOS'
        ws['E4'] = 'DPI'
        cont=5
        #Recorremos el conjunto de personas y vamos escribiendo cada uno de los datos en las celdas
        for persona in personas:
            ws.cell(row=cont,column=2).value = persona.idpersona
            ws.cell(row=cont,column=3).value = persona.nombres_persona
            ws.cell(row=cont,column=4).value = persona.apellidos_persona
            ws.cell(row=cont,column=5).value = persona.dpi_persona
            cont = cont + 1
        #Establecemos el nombre del archivo
        nombre_archivo ="ReportePersonasExcel.xlsx"
        #Definimos que el tipo de respuesta a devolver es un archivo de microsoft excel
        response = HttpResponse(content_type="application/ms-excel")
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response
