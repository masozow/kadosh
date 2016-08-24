#Vista genérica para mostrar resultados
from django.views.generic.base import TemplateView
#Workbook nos permite crear libros en excel
from openpyxl import Workbook
from openpyxl.styles import colors
from openpyxl.styles import Font, Color,Fill,PatternFill
from openpyxl.cell import Cell
from openpyxl.worksheet.properties import WorksheetProperties, PageSetupProperties
#Nos devuelve un objeto resultado, en este caso un archivo de excel

from django.http.response import HttpResponse
from .models import *
from django.db.models import Sum, Count
from django.utils import timezone
import datetime
from django.db import connection
import pytz #para poder hacer la suma de los campos
from django.db.models import Q #para poder usar el operador | que funciona como OR
from django.db.models import F

def ValuesQuerySetToDict(vqs):
    return [item for item in vqs]

def to_dict(instance):
    opts = instance._meta
    data = {}
    for f in opts.concrete_fields + opts.many_to_many:
        if isinstance(f, ManyToManyField):
            if instance.pk is None:
                data[f.name] = []
            else:
                data[f.name] = list(f.value_from_object(instance).values_list('pk', flat=True))
        else:
            data[f.name] = f.value_from_object(instance)
    return data

#Nuestra clase hereda de la vista genérica TemplateView
class ReporteCompras(TemplateView):

    def get(self, request, *args, **kwargs):
        #recibo los datos
        fechaini = self.request.GET.get('fechainicial_precio')
        fechafini = self.request.GET.get('fechafinal_precio')
        box=str(self.request.GET.get('checkbox_vrf'))

        fecha1=str(fechaini) #debe ser la fecha más pequeña
        fecha2=str(fechafini) #debe ser la fecha más grande
        fecha1_split=fecha1.split('/')
        fecha2_split=fecha2.split('/')
        if box=="true":
            box=True
        elif box=="false":
            box=False

        if fecha1 and fecha2:
            fecha1_split=fecha1.split('/')
            fecha2_split=fecha2.split('/')
            compras=Compra.objects.filter(fecha_realizacion_compra__range=(datetime.datetime(int(fecha1_split[2]),int(fecha1_split[1]), int(fecha1_split[0]),0,0,0,tzinfo=pytz.UTC), datetime.datetime(int(fecha2_split[2]),int(fecha2_split[1]), int(fecha2_split[0]),23,59,59,tzinfo=pytz.UTC)),vrf_compra=box)
        else:
            compras=Compra.objects.filter(vrf_compra=box)

        #repo_ventas=ValuesQuerySetToDict(ventas_vendedor)

        #Creamos el libro de trabajo
        wb = Workbook()
        #Definimos como nuestra hoja de trabajo, la hoja activa, por defecto la primera del libro
        ws = wb.active
        #En la celda B1 ponemos el texto 'REPORTE DE PERSONAS'
        ws['B1'] = 'Kadosh'
        ws['B2'] = 'REPORTE DE COMPRAS'
        ws['B3'] = datetime.datetime.now()
        if box:
            ws['B4'] = 'Va a registro fiscal'
        else:
            ws['B4'] = 'No va a registro fiscal'
        #Juntamos las celdas desde la B1 hasta la E1, formando una sola celda
        ws.merge_cells('B1:E1')
        ws.merge_cells('B2:E2')
        ws.merge_cells('B3:E3')
        #Colores de fuente de celda y tama;os de letras
        ft = Font(color=colors.RED)
        b1 = ws['B1']
        b1.font = ft
        ft2 = Font(color=colors.BLUE)
        b2 = ws['B2']
        b2.font = ft2

        fuente_encabezados_tabla=Font(bold=True)
        fondo_encabezados_tabla=PatternFill(fill_type='solid',
                                            #fgColor='FF000000') #e0ebeb
                                            start_color='e0ebeb',
                                            end_color='e0ebeb')
        a8=ws['A5']
        a8.font=fuente_encabezados_tabla
        a8.fill=fondo_encabezados_tabla
        b8=ws['B5']
        b8.font=fuente_encabezados_tabla
        b8.fill=fondo_encabezados_tabla
        c8=ws['C5']
        c8.font=fuente_encabezados_tabla
        c8.fill=fondo_encabezados_tabla
        d8=ws['D5']
        d8.font=fuente_encabezados_tabla
        d8.fill=fondo_encabezados_tabla
        e8=ws['E5']
        e8.font=fuente_encabezados_tabla
        e8.fill=fondo_encabezados_tabla
        f8=ws['F5']
        f8.font=fuente_encabezados_tabla
        f8.fill=fondo_encabezados_tabla
        g8=ws['G5']
        g8.font=fuente_encabezados_tabla
        g8.fill=fondo_encabezados_tabla
        h8=ws['H5']
        h8.font=fuente_encabezados_tabla
        h8.fill=fondo_encabezados_tabla
        i8=ws['I5']
        i8.font=fuente_encabezados_tabla
        i8.fill=fondo_encabezados_tabla
        j8=ws['J5']
        j8.font=fuente_encabezados_tabla
        j8.fill=fondo_encabezados_tabla
        k8=ws['K5']
        k8.font=fuente_encabezados_tabla
        k8.fill=fondo_encabezados_tabla
        l8=ws['L5']
        l8.font=fuente_encabezados_tabla
        l8.fill=fondo_encabezados_tabla


        #Creamos los encabezados desde la celda B3 hasta la E3
        ws['A5'] = 'ID'
        ws['B5'] = 'Proveedor'
        ws['C5'] = 'Casa matriz'
        ws['D5'] = 'Fecha sistema'
        ws['E5'] = 'Fecha compra'
        ws['F5'] = 'Fecha recepción'
        ws['G5'] = 'Factura'
        ws['H5'] = 'Invoice'
        ws['I5'] = 'Recibió'
        ws['J5'] = 'Revisó'
        ws['K5'] = 'No. Guía'
        ws['L5'] = 'Total'


        cont=6

        ws.column_dimensions["D"].width = 20.0
        #Recorremos el conjunto de personas y vamos escribiendo cada uno de los datos en las celdas
        for compra in compras:
            ws.cell(row=cont,column=1).value = compra.pk
            ws.cell(row=cont,column=2).value = str(compra.proveedor_idproveedor)
            ws.cell(row=cont,column=3).value = str(compra.casa_matriz)
            ws.cell(row=cont,column=4).value = compra.fecha_compra
            ws.cell(row=cont,column=5).value = compra.fecha_realizacion_compra
            ws.cell(row=cont,column=6).value = compra.fecha_recepcion_compra
            ws.cell(row=cont,column=7).value = compra.numero_factura
            ws.cell(row=cont,column=8).value = compra.numero_invoice
            ws.cell(row=cont,column=9).value = str(compra.empleado_recibio)
            ws.cell(row=cont,column=10).value = str(compra.empleado_reviso)
            ws.cell(row=cont,column=11).value = compra.numero_guia
            ws.cell(row=cont,column=12).value = compra.total_compra
            cont = cont + 1

        #Establecemos el nombre del archivo
        nombre_archivo ="ReporteCompras.xlsx"
        #Definimos que el tipo de respuesta a devolver es un archivo de microsoft excel
        response = HttpResponse(content_type="application/ms-excel")
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response
