#Vista genérica para mostrar resultados
from django.views.generic.base import TemplateView
#Workbook nos permite crear libros en excel
from openpyxl import Workbook
from openpyxl.styles import colors
from openpyxl.styles import Font, Color,Fill,PatternFill
from openpyxl.cell import Cell
from openpyxl.worksheet.properties import WorksheetProperties, PageSetupProperties
#Nos devuelve un objeto resultado, en este caso un archivo de excel
from django.http.response import HttpResponse
from .models import *
from django.db.models import Sum
from django.db.models import Q
from datetime import datetime,timedelta
import pytz #para poder hacer la suma de los campos

#El siguiente método convierte el resultado de "values" en un diccionario
def ValuesQuerySetToDict(vqs):
    return [item for item in vqs]
#Nuestra clase hereda de la vista genérica TemplateView

class ReporteDevolucion(TemplateView):
    def get(self, request, *args, **kwargs):
        #recibo los datos
        fecha1= self.request.GET.get('fechainicial')
        fecha2= self.request.GET.get('fechafinal')
        if fecha1:
            fechasplit1=str(fecha1).split('/')
            dia=fechasplit1[0]
            mes=fechasplit1[1]
            anio=fechasplit1[2]
            fecha1=datetime(int(anio),int(mes), int(dia),0,0,0,tzinfo=pytz.UTC)
            fecha1=fecha1+timedelta(hours=6)
        else:
            fecha1=''

        if fecha2:
            fechasplit2=str(fecha2).split('/')
            dia=fechasplit2[0]
            mes=fechasplit2[1]
            anio=fechasplit2[2]
            fecha2=datetime(int(anio),int(mes), int(dia),23,59,59,tzinfo=pytz.UTC)
            fecha2=fecha2+timedelta(hours=6)
        else:
            fecha2=''

        if not fecha1 and not fecha2:
            resultado_devolucion=Devolucion.objects.all()
        else:
		   #Cuando existe por lo menos un parámetro (los números dentro de filter deben reemplazarse por los parámetros que envía el usuario)
           resultado_devolucion=Devolucion.objects.filter(momento_devolucion__range=(fecha1,fecha2))
        #if not resultado_cierre:

		   #Cuando existe por lo menos un parámetro (los números dentro de filter deben reemplazarse por los parámetros que envía el usuario)

        #nuevos_productos=ValuesQuerySetToDict(productos)
        #Creamos el libro de trabajo
        wb = Workbook()
        #Definimos como nuestra hoja de trabajo, la hoja activa, por defecto la primera del libro
        ws = wb.active
        #En la celda B1 ponemos el texto 'REPORTE DE PERSONAS'
        ws['B1'] = 'Kadosh'
        ws['B2'] = 'REPORTE DE DEVOLUCIÓN'
        import datetime
        ws['B3'] = datetime.datetime.now()

        #Juntamos las celdas desde la B1 hasta la E1, formando una sola celda
        ws.merge_cells('B1:E1')
        ws.merge_cells('B2:E2')
        ws.merge_cells('B3:E3')
        #Colores de fuente de celda y tama;os de letras
        ft = Font(color=colors.RED)
        b1 = ws['B1']
        b1.font = ft
        ft2 = Font(color=colors.BLUE)
        b2 = ws['B2']
        b2.font = ft2

        #Creamos los encabezados desde la celda B3 hasta la E3
        ws['A5'] = 'Cod'
        ws['B5'] = 'Motivo'
        ws['C5'] = 'Fecha'
        ws['D5'] = 'Venta'

        cont=6
        #Recorremos el conjunto de personas y vamos escribiendo cada uno de los datos en las celdas
        for res in resultado_devolucion:
            ws.cell(row=cont,column=1).value = res.pk
            ws.cell(row=cont,column=2).value = res.motivo_devolucion
            ws.cell(row=cont,column=3).value = res.momento_devolucion
            ws.cell(row=cont,column=4).value = str(res.venta_idventa.pk)
            cont = cont + 1
        #Establecemos el nombre del archivo
        nombre_archivo ="ReporteDevoluciones.xlsx"
        #Definimos que el tipo de respuesta a devolver es un archivo de microsoft excel
        response = HttpResponse(content_type="application/ms-excel")
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response
