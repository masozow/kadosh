#Vista genérica para mostrar resultados
from django.views.generic.base import TemplateView
#Workbook nos permite crear libros en excel
from openpyxl import Workbook
from openpyxl.styles import colors
from openpyxl.styles import Font, Color,Fill,PatternFill
from openpyxl.cell import Cell
from openpyxl.worksheet.properties import WorksheetProperties, PageSetupProperties
#Nos devuelve un objeto resultado, en este caso un archivo de excel
from django.http.response import HttpResponse
from .models import *
from django.db.models import Sum
from django.db.models import Q
from datetime import timedelta
#import datetime
import pytz

#El siguiente método convierte el resultado de "values" en un diccionario
def ValuesQuerySetToDict(vqs):
    return [item for item in vqs]
#Nuestra clase hereda de la vista genérica TemplateView

class ReporteGastos(TemplateView):
    def get(self, request, *args, **kwargs):
        #recibo los datos
        caja= self.request.GET.get('caja_idcaja')
        fecha= self.request.GET.get('momento_gasto')
        import datetime
        if fecha:
            fecha1_split=str(fecha).split('/')
            fechainicial_real=datetime.datetime(int(fecha1_split[2]),int(fecha1_split[1]), int(fecha1_split[0]),0,0,0,tzinfo=pytz.UTC)
            fechafinal_real=datetime.datetime(int(fecha1_split[2]),int(fecha1_split[1]), int(fecha1_split[0]),23,59,59,tzinfo=pytz.UTC)
            fechainicial_real=fechainicial_real+datetime.timedelta(hours=6)
            fechafinal_real=fechafinal_real+datetime.timedelta(hours=6)
         #el not indica que la cadena está vacía, o que la variable es null
        if not caja:
            caja=0

        if not fecha:
            resultado_gastos=Gastos.objects.filter(caja_idcaja=caja)
        else:
            if caja:
                resultado_gastos=Gastos.objects.filter(caja_idcaja=caja,momento_gasto__range=(fechainicial_real,fechafinal_real))
            elif fecha:
                resultado_gastos=Gastos.objects.filter(momento_gasto__range=(fechainicial_real,fechafinal_real))
            else:
                resultado_gastos=Gastos.objects.all()
        #if not resultado_cierre:

		   #Cuando existe por lo menos un parámetro (los números dentro de filter deben reemplazarse por los parámetros que envía el usuario)

        #nuevos_productos=ValuesQuerySetToDict(productos)
        #Creamos el libro de trabajo
        wb = Workbook()
        #Definimos como nuestra hoja de trabajo, la hoja activa, por defecto la primera del libro
        ws = wb.active
        #En la celda B1 ponemos el texto 'REPORTE DE PERSONAS'
        ws['B1'] = 'Kadosh'
        ws['B2'] = 'REPORTE DE CIERRE GASTOS'
        import datetime
        ws['B3'] = datetime.datetime.now()

        #Juntamos las celdas desde la B1 hasta la E1, formando una sola celda
        ws.merge_cells('B1:E1')
        ws.merge_cells('B2:E2')
        ws.merge_cells('B3:E3')
        #Colores de fuente de celda y tama;os de letras
        ft = Font(color=colors.RED)
        b1 = ws['B1']
        b1.font = ft
        ft2 = Font(color=colors.BLUE)
        b2 = ws['B2']
        b2.font = ft2

        #Creamos los encabezados desde la celda B3 hasta la E3
        ws['A5'] = 'Cod'
        ws['B5'] = 'Caja'
        ws['C5'] = 'Fecha'
        ws['D5'] = 'Motivo'

        cont=6
        #Recorremos el conjunto de personas y vamos escribiendo cada uno de los datos en las celdas
        for res in resultado_gastos:
            ws.cell(row=cont,column=1).value = res.pk
            ws.cell(row=cont,column=2).value = str(res.caja_idcaja)
            ws.cell(row=cont,column=3).value = res.momento_gasto
            ws.cell(row=cont,column=4).value = res.motivo_gasto

            cont = cont + 1
        #Establecemos el nombre del archivo
        nombre_archivo ="ReporteGastos.xlsx"
        #Definimos que el tipo de respuesta a devolver es un archivo de microsoft excel
        response = HttpResponse(content_type="application/ms-excel")
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response
