#Vista genérica para mostrar resultados
from django.views.generic.base import TemplateView
#Workbook nos permite crear libros en excel
from openpyxl import Workbook
from openpyxl.styles import colors
from openpyxl.styles import Font, Color,Fill,PatternFill
from openpyxl.cell import Cell
from openpyxl.worksheet.properties import WorksheetProperties, PageSetupProperties
#Nos devuelve un objeto resultado, en este caso un archivo de excel
from django.http.response import HttpResponse
from .models import *
from django.db.models import Sum
from django.db.models import Q
#El siguiente método convierte el resultado de "values" en un diccionario
def ValuesQuerySetToDict(vqs):
    return [item for item in vqs]
#Nuestra clase hereda de la vista genérica TemplateView

class ReporteProductos(TemplateView):
    def get(self, request, *args, **kwargs):
        #recibo los datos
        resp_producto=Producto.objects.filter(inventarioproducto__detalleventa__venta_idventa__contado_venta=1,inventarioproducto__detalleventa__venta_idventa__es_cotizacion=0,inventarioproducto__detalleventa__venta_idventa__estado_venta=1).values('pk','nombre_producto','codigobarras_producto','codigoestilo_producto','marca_id_marca__nombre_marca','tipo_producto_idtipo_producto__nombre_tipoproducto','talla_idtalla__nombre_talla','color_idcolor__nombre_color','genero_idgener__nombre_genero','estilo_idestilo__nombre_estilo').annotate(cantidad_vendida=Sum('inventarioproducto__detalleventa__cantidad_venta'),total_ventas=Sum('inventarioproducto__detalleventa__valor_parcial_venta'))
        txt_codigo_barras = request.GET.get('codigobarras_producto')
        if txt_codigo_barras:
            resp_producto=resp_producto.filter(codigobarras_producto=txt_codigo_barras)
        else:
            txt_codigo_producto = request.GET.get('codigoestilo_producto')
            if txt_codigo_producto:
                resp_producto=resp_producto.filter(codigoestilo_producto=txt_codigo_producto)
            
            id_marca_producto = request.GET.get('marca_id_marca')
            if id_marca_producto:
                resp_producto=resp_producto.filter(marca_id_marca=id_marca_producto)

            id_estilo_producto = request.GET.get('estilo_idestilo')
            if id_estilo_producto:
                resp_producto=resp_producto.filter(estilo_idestilo=id_estilo_producto)

            id_tipo_producto = request.GET.get('tipo_producto_idtipo_producto')
            if id_tipo_producto:
                resp_producto=resp_producto.filter(tipo_producto_idtipo_producto=id_tipo_producto)

            id_talla_producto = request.GET.get('talla_idtalla')
            if id_talla_producto:
                resp_producto=resp_producto.filter(talla_idtalla=id_talla_producto)

            id_color_producto = request.GET.get('color_idcolor')
            if id_color_producto:
                resp_producto=resp_producto.filter(color_idcolor=id_color_producto)

            id_genero_producto = request.GET.get('genero_idgener')
            if id_genero_producto:
                resp_producto=resp_producto.filter(genero_idgener=id_genero_producto)
        
        nuevos_productos=ValuesQuerySetToDict(resp_producto)
        #Creamos el libro de trabajo
        wb = Workbook()
        #Definimos como nuestra hoja de trabajo, la hoja activa, por defecto la primera del libro
        ws = wb.active
        #En la celda B1 ponemos el texto 'REPORTE DE PERSONAS'
        ws['B1'] = 'Kadosh'
        ws['B2'] = 'REPORTE DE PRODUCTOS'
        import datetime
        ws['B3'] = datetime.datetime.now()

        #Juntamos las celdas desde la B1 hasta la E1, formando una sola celda
        ws.merge_cells('B1:E1')
        ws.merge_cells('B2:E2')
        ws.merge_cells('B3:E3')
        #Colores de fuente de celda y tama;os de letras
        ft = Font(color=colors.RED)
        b1 = ws['B1']
        b1.font = ft
        ft2 = Font(color=colors.BLUE)
        b2 = ws['B2']
        b2.font = ft2

        #Creamos los encabezados desde la celda B3 hasta la E3
        ws['B5'] = 'Codigo'
        ws['C5'] = 'Nombre Producto'
        ws['D5'] = 'Codigo Barra'
        ws['E5'] = 'Codigo Estilo'
        #ws['F5'] = 'Precio'
        ws['F5'] = 'Marca'
        ws['G5'] = 'Tipo'
        ws['H5'] = 'Talla'
        ws['I5'] = 'Color'
        ws['J5'] = 'Genero'
        ws['K5'] = 'Estilo'
        ws['L5'] = 'Cantidad Vendida'
        ws['M5'] = 'Total De Ventas'

        cont=6
        #Recorremos el conjunto de personas y vamos escribiendo cada uno de los datos en las celdas
        for producto in nuevos_productos:
            ws.cell(row=cont,column=2).value = producto['pk']
            ws.cell(row=cont,column=3).value = producto['nombre_producto']
            ws.cell(row=cont,column=4).value = producto['codigobarras_producto']
            ws.cell(row=cont,column=5).value = producto['codigoestilo_producto']
            #ws.cell(row=cont,column=6).value = producto['precio__valor_precio']
            ws.cell(row=cont,column=6).value = producto['marca_id_marca__nombre_marca']
            ws.cell(row=cont,column=7).value = producto['tipo_producto_idtipo_producto__nombre_tipoproducto']
            ws.cell(row=cont,column=8).value = producto['talla_idtalla__nombre_talla']
            ws.cell(row=cont,column=9).value = producto['color_idcolor__nombre_color']
            ws.cell(row=cont,column=10).value = producto['genero_idgener__nombre_genero']
            ws.cell(row=cont,column=11).value = producto['estilo_idestilo__nombre_estilo']
            ws.cell(row=cont,column=12).value = producto['cantidad_vendida']
            ws.cell(row=cont,column=13).value = producto['total_ventas']
            cont = cont + 1
        #Establecemos el nombre del archivo
        nombre_archivo ="ReporteProductos.xlsx"
        #Definimos que el tipo de respuesta a devolver es un archivo de microsoft excel
        response = HttpResponse(content_type="application/ms-excel")
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response
