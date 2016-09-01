#Vista genérica para mostrar resultados
from django.views.generic.base import TemplateView
#Workbook nos permite crear libros en excel
from openpyxl import Workbook
from openpyxl.styles import colors
from openpyxl.styles import Font, Color,Fill,PatternFill
from openpyxl.cell import Cell
from openpyxl.worksheet.properties import WorksheetProperties, PageSetupProperties
#Nos devuelve un objeto resultado, en este caso un archivo de excel
from django.http.response import HttpResponse
from .models import *
from django.db.models import Sum
from django.db.models import Q
#El siguiente método convierte el resultado de "values" en un diccionario
def ValuesQuerySetToDict(vqs):
    return [item for item in vqs]
#Nuestra clase hereda de la vista genérica TemplateView

class ReporteProductos(TemplateView):
    def get(self, request, *args, **kwargs):
        #recibo los datos
        caja= self.request.GET.get('caja_idcaja')
        fecha= self.request.GET.get('fecha_cierredecaja')
        if fecha:
            fechasplit=str(fecha).split('/')
            dia=fechasplit[0]
            mes=fechasplit[1]
            anio=fechasplit[2]
            fecha=anio+'-'+mes+'-'+dia
        else:
            fecha=''
         #el not indica que la cadena está vacía, o que la variable es null
        if not caja:
            caja=0

        if not fecha and not caja:
            resultado_cierre=CierreDeCaja.objects.all()
        else:
            resultado_cierre=CierreDeCaja.objects.filter(Q(caja_idcaja=caja)|Q(fecha_cierredecaja=fecha))
        #if not resultado_cierre:

		   #Cuando existe por lo menos un parámetro (los números dentro de filter deben reemplazarse por los parámetros que envía el usuario)

        #nuevos_productos=ValuesQuerySetToDict(productos)
        #Creamos el libro de trabajo
        wb = Workbook()
        #Definimos como nuestra hoja de trabajo, la hoja activa, por defecto la primera del libro
        ws = wb.active
        #En la celda B1 ponemos el texto 'REPORTE DE PERSONAS'
        ws['B1'] = 'Kadosh'
        ws['B2'] = 'REPORTE DE CIERRE DE CAJA'
        import datetime
        ws['B3'] = datetime.datetime.now()

        #Juntamos las celdas desde la B1 hasta la E1, formando una sola celda
        ws.merge_cells('B1:E1')
        ws.merge_cells('B2:E2')
        ws.merge_cells('B3:E3')
        #Colores de fuente de celda y tama;os de letras
        ft = Font(color=colors.RED)
        b1 = ws['B1']
        b1.font = ft
        ft2 = Font(color=colors.BLUE)
        b2 = ws['B2']
        b2.font = ft2

        #Creamos los encabezados desde la celda B3 hasta la E3
        ws['A5'] = 'Cod'
        ws['B5'] = 'Caja'
        ws['C5'] = 'Empleado autorizó'
        ws['D5'] = 'Fecha cierre'
        #ws['F5'] = 'Precio'
        ws['E5'] = 'Total contabilizado'
        ws['F5'] = 'Total(ingresos-egresos)'
        ws['G5'] = 'Finalizado'
        ws['H5'] = 'Total efectivo'
        ws['I5'] = 'Total cheque'
        ws['J5'] = 'Total tarjeta'
        ws['K5'] = 'Total gastos'

        cont=6
        #Recorremos el conjunto de personas y vamos escribiendo cada uno de los datos en las celdas
        for res in resultado_cierre:
            ws.cell(row=cont,column=1).value = res.pk
            ws.cell(row=cont,column=2).value = str(res.caja_idcaja)
            ws.cell(row=cont,column=3).value = str(res.empleado_idempleado)
            ws.cell(row=cont,column=4).value = res.fecha_cierredecaja
            #ws.cell(row=cont,column=6).value = producto['precio__valor_precio']
            ws.cell(row=cont,column=5).value = res.total_real_cierredecaja
            ws.cell(row=cont,column=6).value = res.total_calculado_cierredecaja
            ws.cell(row=cont,column=7).value = res.finalizado_cierredecaja
            ws.cell(row=cont,column=8).value = res.total_efectivo_cierredecaja
            ws.cell(row=cont,column=9).value = res.total_cheque_cierredecaja
            ws.cell(row=cont,column=10).value = res.total_tarjeta_cierredecaja
            ws.cell(row=cont,column=11).value = res.total_egresos_cierredecaja
            cont = cont + 1
        #Establecemos el nombre del archivo
        nombre_archivo ="ReporteCierreDeCaja.xlsx"
        #Definimos que el tipo de respuesta a devolver es un archivo de microsoft excel
        response = HttpResponse(content_type="application/ms-excel")
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response
