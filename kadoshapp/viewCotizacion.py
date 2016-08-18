#Vista genérica para mostrar resultados
from django.views.generic.base import TemplateView
#Workbook nos permite crear libros en excel
from openpyxl import Workbook
from openpyxl.styles import colors
from openpyxl.styles import Font, Color,Fill,PatternFill
from openpyxl.worksheet.properties import WorksheetProperties, PageSetupProperties
from openpyxl.drawing.image import Image
from django.conf import settings

#Nos devuelve un objeto resultado, en este caso un archivo de excel
from django.http.response import HttpResponse
#from django.views.generic.list import ListView
from .models import *

#El siguiente método convierte el resultado de "values" en un diccionario
def ValuesQuerySetToDict(vqs):
    return [item for item in vqs]

#Nuestra clase hereda de la vista genérica TemplateView
class ReporteCotizacion(TemplateView):

    #Usamos el método get para generar el archivo excel
    def get(self, request, *args, **kwargs):
        variable= self.request.GET.get('numero_de_venta')
        #context = self.get_context_data()
        #valor=context["numero_de_venta"]
        #Creamos el libro de trabajo
        wb = Workbook()
        #Definimos como nuestra hoja de trabajo, la hoja activa, por defecto la primera del libro
        ws = wb.active

        detalle=DetalleVenta.objects.filter(venta_idventa=53,inventario_producto_idinventario_producto__producto_codigo_producto__productohasfotografia__fotografia_idfotografia__principal_fotografia=1).values('inventario_producto_idinventario_producto__producto_codigo_producto',
                                                                        'cantidad_venta',
                                                                        'inventario_producto_idinventario_producto__producto_codigo_producto__nombre_producto',
                                                                        'inventario_producto_idinventario_producto__producto_codigo_producto__marca_id_marca__nombre_marca',
                                                                        'inventario_producto_idinventario_producto__producto_codigo_producto__codigoestilo_producto',
                                                                        'inventario_producto_idinventario_producto__producto_codigo_producto__talla_idtalla__nombre_talla',
                                                                        'inventario_producto_idinventario_producto__producto_codigo_producto__color_idcolor__nombre_color',
                                                                        'inventario_producto_idinventario_producto__producto_codigo_producto__genero_idgener__nombre_genero',
                                                                        'inventario_producto_idinventario_producto__producto_codigo_producto__codigoestilo_producto',
                                                                        'valor_parcial_venta')
                                                                        #'inventario_producto_idinventario_producto__producto_codigo_producto__productohasfotografia__fotografia_idfotografia__ruta_fotografia')
        detalle_diccionario=ValuesQuerySetToDict(detalle)
        venta=Venta.objects.filter(idventa=variable).values('cliente_idcliente__persona_idpersona__nombres_persona',
                                                            'cliente_idcliente__persona_idpersona__apellidos_persona',
                                                            'cliente_idcliente__nit_cliente',
                                                            'vendedor_venta__persona_idpersona__nombres_persona',
                                                            'vendedor_venta__persona_idpersona__apellidos_persona',
                                                            'empleado_idempleado__persona_idpersona__nombres_persona',
                                                            'empleado_idempleado__persona_idpersona__apellidos_persona',
                                                            'fecha_venta',
                                                            'total_venta')
        venta_diccionario=ValuesQuerySetToDict(venta)

        venta_simple=venta_diccionario[0]
        ws['A1'] = 'Kadosh'
        ws['A3'] = 'Cotización'
        ws['A4'] = 'Total:'

        ws['E1'] = 'Fecha:'
        ws['E2'] = 'Nit cliente:'
        ws['E3'] = 'Cliente:'
        ws['E4'] = 'Vendedor:'
        ws['E5'] = 'Empleado:'

        ws['C4'] = venta_simple['total_venta']
        ws['F1'] = venta_simple['fecha_venta']
        ws['F2'] = venta_simple['cliente_idcliente__persona_idpersona__nombres_persona']+' '+venta_simple['cliente_idcliente__persona_idpersona__apellidos_persona']
        ws['F3'] = venta_simple['cliente_idcliente__nit_cliente']
        ws['F4'] = venta_simple['vendedor_venta__persona_idpersona__nombres_persona']+' '+venta_simple['vendedor_venta__persona_idpersona__apellidos_persona']
        ws['F5'] = venta_simple['empleado_idempleado__persona_idpersona__nombres_persona']+' '+venta_simple['empleado_idempleado__persona_idpersona__apellidos_persona']
        #Juntamos las celdas, formando una sola celda
        ws.merge_cells('A1:C2')
        ws.merge_cells('A3:C3')
        ws.merge_cells('A4:B4')

        ws.merge_cells('F1:G1')
        ws.merge_cells('F2:H2')
        ws.merge_cells('F3:H3')
        ws.merge_cells('F4:H4')
        ws.merge_cells('F5:H5')
        #Colores de fuente de celda
        ft = Font(color=colors.BLUE,size=20)
        a1 = ws['A1']
        a1.font = ft

        fuente_encabezados_tabla=Font(bold=True)
        fondo_encabezados_tabla=PatternFill(fill_type='solid',
                                            #fgColor='FF000000') #e0ebeb
                                            start_color='e0ebeb',
                                            end_color='e0ebeb')
        a8=ws['A8']
        a8.font=fuente_encabezados_tabla
        a8.fill=fondo_encabezados_tabla
        b8=ws['B8']
        b8.font=fuente_encabezados_tabla
        b8.fill=fondo_encabezados_tabla
        c8=ws['C8']
        c8.font=fuente_encabezados_tabla
        c8.fill=fondo_encabezados_tabla
        d8=ws['D8']
        d8.font=fuente_encabezados_tabla
        d8.fill=fondo_encabezados_tabla
        e8=ws['E8']
        e8.font=fuente_encabezados_tabla
        e8.fill=fondo_encabezados_tabla
        f8=ws['F8']
        f8.font=fuente_encabezados_tabla
        f8.fill=fondo_encabezados_tabla
        g8=ws['G8']
        g8.font=fuente_encabezados_tabla
        g8.fill=fondo_encabezados_tabla
        h8=ws['H8']
        h8.font=fuente_encabezados_tabla
        h8.fill=fondo_encabezados_tabla

        #Creamos los encabezados desde la celda B3 hasta la E3
        ws['A8'] = 'Cod'
        ws['B8'] = 'Cant'
        ws['C8'] = 'Producto'
        ws['D8'] = 'Marca'
        ws['E8'] = 'Color'
        ws['F8'] = 'Talla'
        ws['G8'] = 'Cod.Estilo'
        ws['H8'] ='Valor parcial'
        cont=9
        ws.column_dimensions["A"].width = 4.0
        ws.column_dimensions["B"].width = 5.0
        ws.column_dimensions["C"].width = 20.0
        ws.column_dimensions["D"].width = 7.0
        ws.column_dimensions["F"].width = 5.0
        ws.column_dimensions["H"].width = 13.0
        #############probando insertar imagene
        #img = Image('/archivos/blusapolo.jpg')#,size=(75,100)
        #img.anchor(ws.cell('J1'))
        #ws.add_image(img)

        #Recorremos el conjunto de personas y vamos escribiendo cada uno de los datos en las celdas
        for det in detalle_diccionario:
            ws.cell(row=cont,column=1).value = det['inventario_producto_idinventario_producto__producto_codigo_producto']
            ws.cell(row=cont,column=2).value = det['cantidad_venta']
            ws.cell(row=cont,column=3).value = det['inventario_producto_idinventario_producto__producto_codigo_producto__genero_idgener__nombre_genero']+'-'+det['inventario_producto_idinventario_producto__producto_codigo_producto__nombre_producto']
            ws.cell(row=cont,column=4).value = det['inventario_producto_idinventario_producto__producto_codigo_producto__marca_id_marca__nombre_marca']
            ws.cell(row=cont,column=5).value = det['inventario_producto_idinventario_producto__producto_codigo_producto__color_idcolor__nombre_color']
            ws.cell(row=cont,column=6).value = det['inventario_producto_idinventario_producto__producto_codigo_producto__talla_idtalla__nombre_talla']
            ws.cell(row=cont,column=7).value = det['inventario_producto_idinventario_producto__producto_codigo_producto__codigoestilo_producto']
            ws.cell(row=cont,column=8).value = det['valor_parcial_venta']
            cont = cont + 1
        #Establecemos el nombre del archivo
        nombre_archivo ="CotizacionProductos.xlsx"
        #Definimos que el tipo de respuesta a devolver es un archivo de microsoft excel
        response = HttpResponse(content_type="application/ms-excel")
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response
